import requests
import os
from decimal import Decimal
from datetime import datetime
import pytz
import secrets
import platform
import urllib3
import openpyxl
from wsimple.api import Wsimple

TWOPLACES = Decimal(10) ** -2

# hide non https ssl cert warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# API endpoints
if platform.system() == 'Darwin':
    ib_api_url = "https://localhost:5001/v1/api"
else:
    ib_api_url = "https://localhost:5000/v1/api"


# Questrade rotates the refresh token on every refresh (each one is single-use),
# so we persist the latest refresh token between runs in this file. Override the
# location by setting QT_REFRESH_TOKEN_PATH in secrets.py.
QT_REFRESH_TOKEN_PATH = getattr(
    secrets,
    "QT_REFRESH_TOKEN_PATH",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "qt_refresh_token.txt"),
)

# Token endpoint. Practice/sandbox accounts must redeem tokens against
# https://practicelogin.questrade.com/oauth2/token instead; override via secrets.py.
QT_TOKEN_URL = getattr(
    secrets, "QT_TOKEN_URL", "https://login.questrade.com/oauth2/token"
)

# Wealthsimple fetching is on by default. Set ENABLE_WEALTHSIMPLE = False in
# secrets.py to skip it on a specific machine (e.g. one without WS accounts).
ENABLE_WEALTHSIMPLE = getattr(secrets, "ENABLE_WEALTHSIMPLE", True)


def read_qt_refresh_token():
    try:
        with open(QT_REFRESH_TOKEN_PATH, "r") as token_file:
            return token_file.read().strip() or None
    except FileNotFoundError:
        return None


def save_qt_refresh_token(refresh_token):
    # Write atomically (temp file + os.replace) so an interrupted run can't leave a
    # half-written token that would lock us out and force registering yet another
    # device. Create the file 0600 so the credential isn't world-readable under a
    # typical 022 umask, then re-assert 0600 on the final file.
    tmp_path = QT_REFRESH_TOKEN_PATH + ".tmp"
    fd = os.open(tmp_path, os.O_WRONLY | os.O_CREAT | os.O_TRUNC, 0o600)
    with os.fdopen(fd, "w") as token_file:
        token_file.write(refresh_token.strip())
    os.replace(tmp_path, QT_REFRESH_TOKEN_PATH)
    try:
        os.chmod(QT_REFRESH_TOKEN_PATH, 0o600)
    except OSError:
        # Best effort: some platforms/filesystems don't support POSIX permissions.
        pass


def refresh_qt_access_token(refresh_token):
    # Exchange the stored refresh token for a fresh access token. Using this grant
    # (instead of the interactive authorize flow) reuses a single device
    # authorization instead of registering a new one on every run. POST with a form
    # body keeps the token out of the request URL (and out of any logged URL).
    data = {"grant_type": "refresh_token", "refresh_token": refresh_token}
    response = requests.post(QT_TOKEN_URL, data=data)
    if response.status_code != 200:
        raise Exception(
            f"Questrade token refresh failed ({response.status_code}): {response.text}. "
            f"The refresh token may be expired or already used - generate a new one from "
            f"the Questrade 'My apps' page and save it to {QT_REFRESH_TOKEN_PATH}."
        )
    return response.json()


def qt_endpoint(api_server, path):
    # Questrade's api_server normally ends in "/", but some token responses include
    # the "/v1" version segment. Normalize both so we never build ".../v1v1/...".
    base = api_server.rstrip("/")
    if base.endswith("/v1"):
        base = base[: -len("/v1")]
    return f"{base}/v1/{path}"


def fetch_qt_portfolio_accounts(access_token, api_server):
    headers = {
        "Authorization": f"Bearer {access_token}"
    }
    response = requests.get(qt_endpoint(api_server, "accounts"), headers=headers)
    return response.json()


def fetch_ib_portfolio_accounts():
    response = requests.get(f"{ib_api_url}/portfolio/accounts", verify=False)
    if response.status_code != 200:
        raise Exception(f"API Error {response.status_code}: {response.text}")
    return response.json()

def fetch_interactive_brokers_data():
    try:
        print("Fetching Interactive Brokers account balances...\n")
        ib_portfolio_accounts = fetch_ib_portfolio_accounts()
        for account in ib_portfolio_accounts:
            resp = requests.get(f"{ib_api_url}/portfolio/{account['accountId']}/allocation", verify=False)
            if resp.status_code != 200:
                raise Exception(f"API Error fetching allocation for {account['accountId']}: {resp.status_code} {resp.text}")
            data = resp.json()

            stk_value = data.get('assetClass', {}).get('long', {}).get('STK', 0)
            cash_value = data.get('assetClass', {}).get('long', {}).get('CASH', 0)
            opt_value = data.get('assetClass', {}).get('long', {}).get('OPT', 0)
            fut_value = data.get('assetClass', {}).get('long', {}).get('FUT', 0)

            # Calculate the formatted total
            total_value = stk_value + cash_value + opt_value + fut_value

            # Account description setup
            account_desc = account['desc']
            if account_desc == secrets.IBKR_CASH_ACCOUNT_ID:
                account_desc = 'IBKR (CASH)'
            if account_desc == secrets.IBKR_TFSA_ACCOUNT_ID:
                account_desc = 'IBKR (TFSA)'
            if account_desc == secrets.IBKR_RRSP_ACCOUNT_ID:
                account_desc = 'IBKR (RRSP)'

            # Call the update_spreadsheet function to update the spreadsheet
            update_spreadsheet(account_desc, total_value)
        print("Interactive Brokers account balances fetched and spreadsheet updated successfully!\n")
    except Exception as e:
        print("An error occurred with IBKR:", e)
        # We don't raise here, we let main continue or handle it gracefully if needed, 
        # or we keep raise so the script stops. But let's check auth.
        raise


def fetch_questrade_data():
    try:
        refresh_token = read_qt_refresh_token()
        if not refresh_token:
            # One-time bootstrap: generate a token from the Questrade "My apps"
            # page (select your app -> generate a new token) and paste it here.
            # Do this only once - from now on the script rotates it automatically.
            print(
                "No saved Questrade refresh token found.\n"
                "Generate one from the Questrade 'My apps' page (select your app -> "
                "generate a new token) and paste it below. You only need to do this once.\n"
            )
            refresh_token = input("Enter your Questrade refresh token: ").strip()
            if not refresh_token:
                raise Exception("No Questrade refresh token provided.")

        print("Refreshing Questrade access token...\n")
        token_data = refresh_qt_access_token(refresh_token)

        # Questrade rotated the refresh token (single-use), so persist the new one
        # immediately, before any API calls. If saving fails, print it so the user
        # can store it manually instead of being stranded with a spent token and
        # forced to register a new device.
        new_refresh_token = token_data["refresh_token"]
        try:
            save_qt_refresh_token(new_refresh_token)
        except OSError as save_error:
            print(
                f"WARNING: could not save the rotated Questrade refresh token to "
                f"{QT_REFRESH_TOKEN_PATH} ({save_error}).\n"
                f"Save this value to that file manually or it will be lost and you'll "
                f"need to generate a new token:\n\n    {new_refresh_token}\n"
            )

        access_token = token_data["access_token"]
        qt_api_url = token_data["api_server"]

        print("Authentication successful.\n")
        print("Fetching Questrade account balances...\n")
        qt_portfolio_accounts = fetch_qt_portfolio_accounts(access_token, qt_api_url)
        for account in qt_portfolio_accounts['accounts']:
            account_number = account['number']
            headers = {
                "Authorization": f"Bearer {access_token}"
            }
            data = requests.get(qt_endpoint(qt_api_url, f"accounts/{account_number}/balances"), headers=headers).json()
            account_desc = 'QT ('+account['type']+')'
            for record in data['combinedBalances']:
                if record['currency'] == 'CAD':
                    # Call the update_spreadsheet function to update the spreadsheet
                    update_spreadsheet(account_desc, record['totalEquity'])
        print("Questrade account balances fetched and spreadsheet updated successfully!\n")

    except Exception as e:
        print("An error occurred with Questrade:", e)
        raise


def get_otp():
    return input("Enter OTP number: \n>>>")


def fetch_wealthsimple_trade_data():
    # use the credentials from the secrets.py file
    # check the current authentication tokens, use the auth.tokens API
    ws = Wsimple(secrets.WEALTHSIMPLE_USERNAME, secrets.WEALTHSIMPLE_PASSWORD, otp_callback=get_otp)

    # always check if Wealthsimple is working (return True if working or an error)
    positions = ws.get_positions()
    accounts = ws.get_accounts()

    positions = positions['results']
    accounts = accounts['results']
    rrsp_cash = 0.00
    tfsa_cash = 0.00

    for account in accounts:
        if account['id'] == secrets.WEALTHIMSPLE_RRSP_ACCOUNT_ID:
            rrsp_cash = Decimal(account['current_balance']['amount']).quantize(TWOPLACES)
        if account['id'] == secrets.WEALTHIMSPLE_TFSA_ACCOUNT_ID:
            tfsa_cash = Decimal(account['current_balance']['amount']).quantize(TWOPLACES)

    rrsp_positions = []
    tfsa_positions = []

    for position in positions:
        if position['account_id'] == secrets.WEALTHIMSPLE_RRSP_ACCOUNT_ID:
            rrsp_positions.append(int(position['quantity']) * Decimal(position['quote']['amount']))
        if position['account_id'] == secrets.WEALTHIMSPLE_TFSA_ACCOUNT_ID:
            tfsa_positions.append(int(position['quantity']) * Decimal(position['quote']['amount']))

    update_spreadsheet('WST (RRSP)', Decimal(sum(rrsp_positions)).quantize(TWOPLACES) + rrsp_cash)
    update_spreadsheet('WST (TFSA)', Decimal(sum(tfsa_positions)).quantize(TWOPLACES) + tfsa_cash)

    print("Wealthsimple Trade account balances fetched and spreadsheet updated successfully!\n")


def update_spreadsheet(account_desc, formatted_total):
    # Load the Excel spreadsheet
    # if macos
    # spreadsheet_path = secrets.ACCOUNT_BALANCE_EXCEL_PATH_MACOS
    # if windows
    # spreadsheet_path = secrets.ACCOUNT_BALANCE_EXCEL_PATH_WINDOWS
    if platform.system() == 'Darwin':
        spreadsheet_path = secrets.ACCOUNT_BALANCE_EXCEL_PATH_MACOS
    else:
        spreadsheet_path = secrets.ACCOUNT_BALANCE_EXCEL_PATH_WINDOWS

    workbook = openpyxl.load_workbook(spreadsheet_path)
    # Select the specific sheet by name
    sheet = workbook['Balances']

    # Find the matching account in the spreadsheet and update the adjacent cell
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True), start=2):
        for col_index, cell_value in enumerate(row):
            if cell_value == account_desc:
                adjacent_col_index = col_index + 2  # Get the column index immediately to the right
                adjacent_col_index_Date = col_index + 3 # get the column index for the date
                adjacent_col_index_Time = col_index + 4 # get the column index for the time

                # Update the cell in the adjacent column
                sheet.cell(row=row_index, column=adjacent_col_index, value=formatted_total)

                # Get current time in UTC
                now_utc = datetime.now(pytz.timezone('UTC'))

                # Convert to Eastern Time
                now_est = now_utc.astimezone(pytz.timezone('US/Eastern'))

                sheet.cell(row=row_index, column=adjacent_col_index_Date, value=now_est.strftime('%Y-%m-%d'))
                sheet.cell(row=row_index, column=adjacent_col_index_Time, value=now_est.strftime('%H:%M:%S'))
                break

    # Save the updated spreadsheet
    workbook.save(spreadsheet_path)


# Main function
def main():
    try:
        fetch_questrade_data()
        fetch_interactive_brokers_data()
        if ENABLE_WEALTHSIMPLE:
            fetch_wealthsimple_trade_data()
        else:
            print("Wealthsimple fetching is disabled (ENABLE_WEALTHSIMPLE = False); skipping.\n")
        print("All account balances fetched and spreadsheet updated successfully!\n")

    except Exception as e:
        print(f"Script execution stopped due to an error: {str(e)}")


if __name__ == "__main__":
    main()
